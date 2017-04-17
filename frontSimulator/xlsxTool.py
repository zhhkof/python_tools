from openpyxl import load_workbook


class Excel:
    def __init__(self, path='information.xlsx'):
        self.path = path

    def loadxlsx(self):
        try:
            self.wb = load_workbook(self.path)
        except Exception as e:
            print(e)

    def get_wmanModDict_from_sheet(self, sheet_name='propaths'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        columnsObj = tuple(ws.columns)
        all_value = {}
        if len(columnsObj) > 0:
            for i in range(0, len(columnsObj)):
                column_value = []
                pid = columnsObj[i][0].value
                for cell in columnsObj[i]:
                    # print(cell.value)
                    if cell.value is None:
                        break
                    column_value.append(cell.value)
                del (column_value[0])
                all_value[pid] = column_value
        return all_value

    def get_datalist_from_sheet2(self, sheet_name='info'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        rowsObj = tuple(ws.rows)
        all_value = []
        keys = [x.value for x in rowsObj[0]]
        if len(rowsObj) >= 2:
            for i in range(1, len(rowsObj)):
                # print(rowdata[i])
                row_value = {}
                for j in range(len(rowsObj[i])):
                    if rowsObj[i][0].value is None:
                        break
                    # print(cell.value)
                    row_value[keys[j]] = str(rowsObj[i][j].value)
                # if len(row_value) > 0:
                all_value.append(row_value)
        return all_value

    def get_datalist_from_sheet(self, sheet_name='info'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        rowsObj = tuple(ws.rows)
        all_value = []
        if len(rowsObj) >= 2:
            for i in range(1, len(rowsObj)):
                # print(rowdata[i])
                row_value = []
                for cell in rowsObj[i]:
                    if rowsObj[i][0].value is None:
                        break
                    # print(cell.value)
                    row_value.append(str(cell.value))
                if len(row_value) > 0:
                    all_value.append(row_value)
        return all_value

    def get_sNdict_from_sheet(self, sheet_name='proSeatNum'):
        ws = self.wb.get_sheet_by_name(sheet_name)
        rowsObj = tuple(ws.rows)
        if len(rowsObj) >= 2:
            all_value = {}
            for i in range(1, len(rowsObj)):
                row_value = []
                pid = str(rowsObj[i][0].value)
                for cell in rowsObj[i]:
                    row_value.append(int(cell.value))
                del (row_value[0])
                all_value[pid] = row_value
            return all_value

    def get_Package_data_from_sheet(self, sheet_name="750packages"):
        ws = self.wb.get_sheet_by_name(sheet_name)
        columnsObj = tuple(ws.columns)
        all_value = {}
        if len(columnsObj) > 0:
            for i in range(0, len(columnsObj)):
                column_value = []
                # package_name = columnsObj[i][0].value
                for cell in columnsObj[i]:
                    # print(cell.value)
                    if cell.value is None:
                        break
                    column_value.append(str(cell.value))
                package_name=column_value[0]+'|'+column_value[1]
                del (column_value[0:2])
                all_value[package_name] = column_value
        return all_value


# cells2=tuple(ws.columns)
# print(len(cells2))
# print(ws['1:2'])
# print(ws['A1:C2'])
# print(ws['A:C'])

# if __name__ == '__main__':
#     e = Excel('information.xlsx')
#     e.loadxlsx()
#     print(e.get_Package_data_from_sheet("750packages"))
